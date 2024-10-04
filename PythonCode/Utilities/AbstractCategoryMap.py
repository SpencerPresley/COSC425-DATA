import os
from openpyxl import Workbook
import sys

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))

sys.path.append(project_root)

from GeneralUtilities.file_ops.file_ops import FileOps
from typing import Tuple
from utilities import Utilities
from enums import AttributeTypes
from warning_manager import WarningManager  # for type hinting


class AbstractCategoryMap:
    def __init__(
        self,
        *,
        utilities_obj: Utilities,
        warning_manager: WarningManager,
        dir_path: str,
    ):
        self.utilities = utilities_obj
        self.warning_manager = warning_manager
        self.dir_path = dir_path
        self.file_ops = FileOps(output_dir=".")
        self.results = self.map_abstract_categories(dir_path=self.dir_path)
        self.file_ops.write_json("for_jensen.json", self.results)

    def map_abstract_categories(self, *, dir_path: str):
        results = {}
        for filename in os.listdir(dir_path):
            file_path = os.path.join(dir_path, filename)
            if not os.path.isfile(file_path):
                self.warning_manager.log_warning(
                    "File Processing", f"File not found: {file_path}"
                )
                continue

            file_content = self.file_ops.read_file(file_path)

            attributes = self.utilities.get_attributes(
                file_content,
                [
                    AttributeTypes.TITLE,
                    AttributeTypes.ABSTRACT,
                    AttributeTypes.WC_PATTERN,
                    AttributeTypes.AUTHOR,
                    AttributeTypes.DEPARTMENT,
                ],
            )

            title = (
                attributes[AttributeTypes.TITLE][1]
                if attributes[AttributeTypes.TITLE][0]
                else None
            )
            abstract = (
                attributes[AttributeTypes.ABSTRACT][1]
                if attributes[AttributeTypes.ABSTRACT][0]
                else None
            )
            categories = (
                attributes[AttributeTypes.WC_PATTERN][1]
                if attributes[AttributeTypes.WC_PATTERN][0]
                else []
            )
            authors = (
                attributes[AttributeTypes.AUTHOR][1]
                if attributes[AttributeTypes.AUTHOR][0]
                else []
            )
            department = (
                attributes[AttributeTypes.DEPARTMENT][1]
                if attributes[AttributeTypes.DEPARTMENT][0]
                else []
            )

            if not title:
                self.warning_manager.log_warning(
                    "Attribute Extraction", f"Title not found for file: {file_path}"
                )
            if not abstract:
                self.warning_manager.log_warning(
                    "Attribute Extraction", f"Abstract not found for file: {file_path}"
                )
            else:
                results[title] = {
                    "abstract": abstract,
                    "categories": categories,
                    "authors": authors,
                    "department": department,
                }

        return results

    def get_results(self):
        return self.results

    def get_taxonomy(self):
        taxonomy_path_json = "../../TextAnalysis/Taxonomy.json"
        try:
            taxonomy = self.file_ops.read_json(taxonomy_path_json)
            return taxonomy
        except FileNotFoundError:
            self.warning_manager.log_warning(
                "Taxonomy", f"Taxonomy file not found: {taxonomy_path_json}"
            )
            return {}

    def get_file_ops_obj(self):
        return self.file_ops

    def write_to_excel(self, results, filename: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "Abstract Categories"

        # Write headers
        headers = [
            "Title",
            "Abstract",
            "Categories",
            "Authors",
            "Department",
            "Upper Level Categories",
            "Mid-Level Categories",
            "Themes",
        ]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        # Write data
        for row, (title, data) in enumerate(results.items(), start=2):
            ws.cell(row=row, column=1, value=title)
            ws.cell(row=row, column=2, value=data.get("abstract", ""))
            ws.cell(row=row, column=3, value=", ".join(data.get("categories", [])))
            ws.cell(row=row, column=4, value=", ".join(data.get("authors", [])))

            # Handle 'department' which might be a string or a list
            department = data.get("department", [])
            if isinstance(department, list):
                department = ", ".join(department)
            ws.cell(row=row, column=5, value=department)

            ws.cell(
                row=row,
                column=6,
                value=", ".join(data.get("Upper Level Categories", [])),
            )
            ws.cell(
                row=row, column=7, value=", ".join(data.get("Mid-Level Categories", []))
            )

            # Ensure all items in 'Themes' are strings
            themes = data.get("Themes", [])
            themes = [str(theme) for theme in themes]
            ws.cell(row=row, column=8, value=", ".join(themes))

        try:
            wb.save(filename)
        except Exception as e:
            self.warning_manager.log_warning(
                "Excel Writing", f"Error writing to Excel file {filename}: {str(e)}"
            )


if __name__ == "__main__":
    abstract_category_map = AbstractCategoryMap(
        Utilities_obj=Utilities(), dir_path="./split_files"
    )

    results = abstract_category_map.get_results()
    taxonomy = abstract_category_map.get_taxonomy()

    for key, value in results.items():
        taxonomy_obj = taxonomy[key]
        results[key]["Upper Level Categories"] = taxonomy_obj["Upper Level Categories"]
        results[key]["Mid-Level Categories"] = taxonomy_obj["Mid-Level Categories"]
        results[key]["Themes"] = taxonomy_obj["Themes"]

    file_ops = abstract_category_map.get_file_ops_obj()
    file_ops.write_json("FINAL_ForJensen.json", results)

    # Write to Excel
    abstract_category_map.write_to_excel(results, "CategoryMapping.xlsx")
